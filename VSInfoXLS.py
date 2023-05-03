import openpyxl
from openpyxl.chart import BarChart, Reference, Series

# Cria um novo arquivo do Excel
workbook = openpyxl.Workbook()

# Seleciona a primeira planilha
sheet = workbook.active

# Define os dados do gráfico
musicas = ('Lib Provisória', 'Sentadão', 'Combatchy', 'Surtada', 'Cheirosa')
indice = [1, 2, 3, 4, 5]
acessos = [1068254, 866216, 763652, 758198, 526324]

# Insere os dados na planilha
for i in range(len(musicas)):
    sheet.cell(row=i+1, column=1).value = musicas[i]
    sheet.cell(row=i+1, column=2).value = acessos[i]

# Define a faixa de dados para o gráfico
data = Reference(sheet, min_col=2, min_row=1, max_row=len(musicas))

# Cria um gráfico de barras
chart = BarChart()

# Adiciona os dados do gráfico
chart.add_data(data)

# Define o título do gráfico
chart.title = 'Ranking do Spotify em 21/12/2012'

# Adiciona o gráfico à planilha
sheet.add_chart(chart, "D2")

# Salva o arquivo do Excel
workbook.save("grafico.xlsx")
